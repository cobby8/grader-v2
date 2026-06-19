# -*- coding: utf-8 -*-
"""주문서(xlsx) 파서 — 선수별 행 / 사이즈 집계 두 양식을 읽어 표준 리스트로 변환.

이 모듈은 **순수 입력 파서**다. engine 코어(compose/Piece/parse_svg 등)를
전혀 import 하지 않으며 수정하지도 않는다(독립 모듈). 결과만 넘긴다.

  parse_order(xlsx_path) -> list[dict]
      각 dict = {"name": 이름, "number": 배번, "size": 상의사이즈, "qty": 수량}
      값은 전부 문자열. 추출 실패한 칸은 빈 문자열("")로 두고 행은 살린다
      (크래시 금지 — 웹에서 빈칸을 붉게 강조해 사람이 수기 입력하는 안전망).

STIZ 표준 주문서는 실데이터 86개 전수 분석 결과 크게 2종이다.
  ① 「선수별 행」(주류, 81/86)
       A=이니셜(이름)  B=배번  C=상의사이즈  D=하의사이즈  F=참고사항
       헤더 텍스트 '이니셜/배 번/사이즈'가 2행 또는 3행(양식 버전차) → 행 고정 금지,
       '이니셜' 키워드로 헤더행을 찾는다. 헤더 아래 부제행(상의/하의)을 건너뛰고
       그 다음 행부터 데이터, 빈 행이나 '배송 정보' 같은 푸터를 만나면 종료.
       → 1선수 = 1행. qty 는 "1" 고정(F열 자유텍스트는 파싱하지 않음).
  ② 「사이즈 집계」(소수, 5/86 — 연습복·슈팅저지·추가주문)
       '이니셜' 헤더 없이 A=사이즈 세로나열, B=수량. 이름·배번이 없다.
       → 기존 grader 의 셀 전체 스캔(세로/가로/표) 규칙을 이식해 best-effort 처리.

1차 범위(결정): ①을 메인으로 완성, ②는 best-effort, 나머지 변형은 부분실패+경고.
사이즈는 상의(C열) 기준(하의 분리는 백로그).
"""
from __future__ import annotations

import re

import openpyxl


# ── openpyxl Custom Properties 버그 방어 패치(기존 grader 규칙 이식) ──────────
# 일부 외부 xlsx 는 Custom Document Property 의 name 이 None 이라 로드 중 터진다.
# (StringProperty.name 은 str 이어야 하는데 NoneType 이라 예외). name=None 을 ""로
# 바꿔주는 몽키패치. 현 데이터셋엔 미발생이지만 외부 파일 대비 무해한 안전망이다.
# try/except 로 감싸 패치 자체가 실패해도 본 모듈 동작에는 영향이 없게 한다.
try:
    from openpyxl.packaging.custom import StringProperty  # type: ignore

    _orig_str_init = StringProperty.__init__

    def _patched_str_init(self, *args, **kwargs):
        # name 키워드가 None 으로 들어오면 빈 문자열로 치환해 예외를 막는다.
        if "name" in kwargs and kwargs["name"] is None:
            kwargs["name"] = ""
        _orig_str_init(self, *args, **kwargs)

    StringProperty.__init__ = _patched_str_init  # type: ignore
except Exception:
    # 패치 실패는 무시(openpyxl 버전에 따라 경로가 다를 수 있음) — 본 동작엔 무해.
    pass


# ── 사이즈 사전 ──────────────────────────────────────────────────────────────
# 길이가 긴 것부터 매칭해야 "5XL"이 "XL"로 잘못 잡히지 않는다(중요).
# 아동 호수(피구·줄넘기 76건)는 별도 정규식으로 처리한다(아래 _normalize_size).
SIZE_KEYWORDS = [
    "5XS", "4XS", "3XS", "2XS",
    "5XL", "4XL", "3XL", "2XL",
    "XS", "XL",
    "S", "M", "L",
]
# 매칭은 항상 길이 내림차순으로 시도한다(예: "2XL"을 "XL"보다 먼저 본다).
_SIZE_BY_LEN = sorted(SIZE_KEYWORDS, key=len, reverse=True)

# 아동 호수: "8호", "10 호" 처럼 1~2자리 숫자 뒤에 '호'가 붙은 경우만 사이즈로 인정.
# 한글 이름 끝글자가 '호'인 경우(예: "장현호")는 앞이 숫자가 아니라 걸러진다.
_HOSU_PATTERN = re.compile(r"^(\d{1,2})\s*호$")

# 데이터 영역 종료를 알리는 푸터 키워드(이 텍스트가 셀에 보이면 데이터 끝으로 판단).
_FOOTER_KEYWORDS = ("배송 정보", "배송정보", "성명", "주소", "연락처")

# 헤더 탐색용 키워드(공백 제거 후 비교).
_HEADER_INITIAL = ("이니셜",)          # 양식①/③ 식별 키워드(이름 열)
_HEADER_NUMBER = ("배번",)             # '배 번' → 공백 제거 시 '배번'
_HEADER_SIZE = ("사이즈",)             # '사이즈'

# 신양식③(STIZ '작성하기' 신폼) 헤더 키워드.
# 양식③은 맨 앞에 '순번' 컬럼이 있고, '비고' 컬럼에 "4장" 식 수량이 적힌다.
_HEADER_SEQ = ("순번",)                # '순\xa0\xa0 번' → nbsp/공백 제거 시 '순번'
_HEADER_NOTE = ("비고",)              # '비  고' → 공백 제거 시 '비고'

# ── 이번 주문 표식 키워드 ────────────────────────────────────────────────────
# STIZ 주문서는 한 xlsx 에 여러 접수분 시트가 누적된다(시트명=날짜 6자리).
# **이번 주문 시트**는 상단(보통 1~2행)에 "주문번호 : <건명>" + "수량 : 상의 NEA"
# 표식이 함께 적혀 있고, 과거 접수분 시트에는 이 표식이 없다(안내문만 있음).
# → 이 두 키워드가 상단에 같이 보이는 시트를 '이번 주문'으로 우선 채택한다.
# (실데이터 86개 전수: 82개는 첫 시트에 표식, 1개는 뒤 시트에 표식, 3개는 표식 없음)
_ORDER_MARK_KEYWORDS = ("주문번호", "수량")
_ORDER_MARK_SCAN_ROWS = 5   # 상단 몇 행까지 표식을 찾을지(헤더는 보통 2~3행).
_ORDER_MARK_SCAN_COLS = 8   # 표식은 보통 A열·D열에 있음(여유 있게 A~H).


def _has_order_mark(rows: list) -> bool:
    """시트 상단에 '주문번호'+'수량' 표식이 함께 있으면 True(=이번 주문 시트).

    과거 접수분 시트는 '제작 리스트를 기재해주세요' 같은 안내문만 있어 False.
    공백을 제거하고 비교해 '주 문 번 호' 같은 변형에도 견딘다.
    """
    parts = []
    for r in range(min(_ORDER_MARK_SCAN_ROWS, len(rows))):
        row = rows[r]
        for c in range(min(_ORDER_MARK_SCAN_COLS, len(row))):
            v = row[c]
            if v is not None:
                parts.append(str(v))
    text = "".join(parts).replace(" ", "")
    return all(kw in text for kw in _ORDER_MARK_KEYWORDS)


def _to_str(value) -> str:
    """셀 값을 안전한 문자열로 변환한다.

    - None / 공백 → ""
    - 숫자: 정수면 "24"(소수점 .0 제거), 그 외 보존. → 배번 "0"/"00"은 문자열로 들어오므로
      그대로 살아남는다(엑셀에서 0/00 은 보통 텍스트 셀이라 float 가 아니다).
    - "24.0" 같은 문자열도 정수로 보이면 "24"로 정리.
    """
    if value is None:
        return ""
    # 숫자형: 정수면 .0 떼고, 아니면 그대로 문자열화.
    if isinstance(value, bool):
        # bool 은 int 의 하위형이라 먼저 걸러 "True/False"가 숫자로 새는 걸 막는다.
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        # 13.0 → "13", 13.5 → "13.5"
        return str(int(value)) if value.is_integer() else str(value)
    text = str(value).strip()
    if not text:
        return ""
    # "24.0" 같은 문자열도 정수로 정리(배번이 수식 결과로 들어온 경우 대비).
    m = re.fullmatch(r"(\d+)\.0+", text)
    if m:
        return m.group(1)
    return text


def _normalize_size(value) -> str:
    """셀 값을 표준 사이즈 키워드로 정규화한다. 실패하면 "" 반환.

    예) "2 XL" → "2XL", "2-XL" → "2XL", "l사이즈" → "L", "10호" → "10호",
        "장현호" → ""(이름이므로 사이즈 아님), 알 수 없는 값 → "".
    """
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""

    # (1) 아동 호수: 1~2자리 숫자 + '호' 만 인정(이름 끝 '호' 오탐 차단).
    m = _HOSU_PATTERN.match(text)
    if m:
        return f"{int(m.group(1))}호"

    # (2) 일반 사이즈: 공백/하이픈을 제거하고 대문자로 통일한 뒤 사전과 비교.
    #     "L 사이즈" / "L사이즈" 처럼 뒤에 '사이즈'가 붙은 경우도 떼어낸다.
    cleaned = text.upper().replace(" ", "").replace("-", "")
    cleaned = re.sub(r"(사이즈|SIZE)$", "", cleaned)
    # 긴 키워드부터 비교해 정확히 일치하는 것만 사이즈로 인정(부분일치 금지).
    for kw in _SIZE_BY_LEN:
        if cleaned == kw:
            return kw
    return ""


def _norm_header(value) -> str:
    """헤더 비교용: 공백 제거 + 문자열화(예: '배 번' → '배번').

    일반 공백(" ")뿐 아니라 **nbsp(\\xa0, non-breaking space)** 도 제거한다.
    신양식③ 헤더는 '순\\xa0\\xa0 번', '\\xa0이니셜', '\\xa0배\\xa0\\xa0 번'처럼
    엑셀에서 칸 맞춤용으로 nbsp 를 끼워 넣는다. nbsp 를 안 지우면
    '\\xa0이니셜' 이 '이니셜' 과 달라 헤더 탐색이 실패한다(양식③ 0행 원인).
    양식①/② 헤더엔 nbsp 가 없으므로 이 보강은 그 양식들엔 무손상이다.
    """
    if value is None:
        return ""
    # \xa0(nbsp) → 일반 공백으로 치환 후, 일반 공백 전부 제거.
    return str(value).replace("\xa0", " ").replace(" ", "").strip()


# ── 양식① 「선수별 행」 파서 ──────────────────────────────────────────────────
def _parse_form1(rows: list) -> list[dict]:
    """행 2차원 리스트에서 양식①(선수별 행)을 추출한다.

    rows: [[셀값, ...], ...] (한 시트의 전체 값, iter_rows(values_only=True) 결과)
    반환: [{name, number, size, qty}] — '이니셜' 헤더를 못 찾으면 빈 리스트.
    """
    # 1) '이니셜' 헤더 셀 위치 탐색(상단 1~12행, A~G열 범위).
    header_r = header_c = None
    for r in range(min(12, len(rows))):
        row = rows[r]
        for c in range(min(7, len(row))):
            if _norm_header(row[c]) in _HEADER_INITIAL:
                header_r, header_c = r, c
                break
        if header_r is not None:
            break
    if header_r is None:
        return []  # 양식①이 아님

    # 2) 헤더행에서 배번/사이즈 열을 식별(보통 이니셜=A, 배번=B, 사이즈=C이지만
    #    키워드로 찾아 양식 변형에 견디게 한다). 못 찾으면 표준 오프셋으로 가정.
    name_col = header_c
    number_col = header_c + 1
    size_col = header_c + 2
    hrow = rows[header_r]
    for c in range(len(hrow)):
        h = _norm_header(hrow[c])
        if h in _HEADER_NUMBER:
            number_col = c
        elif h in _HEADER_SIZE:
            size_col = c  # 사이즈 헤더(C). 부제행에서 상의가 이 열에 옴.

    # 3) 데이터 시작행: 헤더 바로 아래가 부제행(상의/하의)이면 건너뛴다.
    #    부제행 판정: 사이즈 열 또는 그 옆에 '상의'/'하의' 텍스트가 있으면 부제행.
    start_r = header_r + 1
    if start_r < len(rows):
        sub = rows[start_r]
        joined = "".join(_norm_header(v) for v in sub[:6])
        if "상의" in joined or "하의" in joined:
            start_r += 1  # 부제행을 건너뛰고 그 다음 행부터 데이터.

    # 4) 데이터 행 추출: 빈 행 또는 푸터 키워드를 만나면 종료.
    result: list[dict] = []
    for r in range(start_r, len(rows)):
        row = rows[r]

        def cell(idx):
            return row[idx] if idx < len(row) else None

        name = _to_str(cell(name_col))
        number = _to_str(cell(number_col))
        size_raw = cell(size_col)
        size = _normalize_size(size_raw)

        # 종료 조건 (a): 행 어딘가에 푸터 키워드가 있으면 데이터 끝.
        row_text = " ".join(_to_str(v) for v in row[:7])
        if any(k in row_text.replace(" ", "") for k in
               (kw.replace(" ", "") for kw in _FOOTER_KEYWORDS)):
            break

        # 배번이 '배번스러운지' 판정: 숫자(0·00 포함)거나 한두 글자(GK 등 표기) 정도여야
        # 선수 행이다. r20 "스타킹 필드용..." 같은 긴 안내문은 배번이 아니다.
        number_is_valid = bool(re.fullmatch(r"\d{1,4}", number)) or (0 < len(number) <= 5)

        # 종료 조건 (b): 사이즈도 없고 '배번스럽지도' 않으면 데이터 영역이 끝난 것으로 본다.
        #   → 빈 행, 안내 문구 행(긴 텍스트가 한 칸에 들어온 경우 등)에서 멈춘다.
        if not size and not number_is_valid:
            break

        # 유효 데이터 행: 배번 또는 사이즈 중 하나라도 '진짜' 있어야 선수 행으로 인정.
        if not size and not number:
            continue

        result.append({
            "name": name,
            "number": number,
            "size": size,   # 정규화 실패 시 ""(웹에서 강조 → 수기 보정)
            "qty": "1",     # 양식①은 1선수=1벌 → qty 고정.
        })

    return result


# ── 양식② 「사이즈 집계」 파서 (기존 grader 규칙 이식) ────────────────────────
def _parse_form2(rows: list) -> list[dict]:
    """행 2차원 리스트에서 양식②(사이즈 집계)를 추출한다.

    셀 전체를 스캔해 사이즈 키워드 위치를 모으고, 세로형/가로형/표형을 판별해
    인접 셀에서 수량을 읽는다(기존 grader order_parser 규칙 이식).
    반환: [{name:"", number:"", size, qty}] — 수량이 0/빈값이면 그 사이즈는 제외.
    """
    # 1) 모든 셀 스캔 → 사이즈 셀 좌표 수집.
    size_cells = []  # [(r, c, size)]
    for r, row in enumerate(rows):
        for c, val in enumerate(row):
            s = _normalize_size(val)
            if s:
                size_cells.append((r, c, s))
    if not size_cells:
        return []

    # 2) 가로형(한 행에 사이즈 여럿) vs 세로형(한 열에 사이즈 여럿) 판별.
    row_counts: dict[int, int] = {}
    col_counts: dict[int, int] = {}
    for r, c, _ in size_cells:
        row_counts[r] = row_counts.get(r, 0) + 1
        col_counts[c] = col_counts.get(c, 0) + 1
    max_row = max(row_counts.values())
    max_col = max(col_counts.values())

    def qty_of(value) -> str:
        """인접 셀 값을 수량 문자열로. 0/빈값/비숫자는 ""."""
        if value is None:
            return ""
        if isinstance(value, bool):
            return ""
        if isinstance(value, (int, float)):
            n = int(value)
            return str(n) if 0 < n < 100000 else ""
        m = re.match(r"^(\d+)", str(value).strip())
        if m:
            n = int(m.group(1))
            return str(n) if 0 < n < 100000 else ""
        return ""

    def get(r, c):
        return rows[r][c] if (0 <= r < len(rows) and 0 <= c < len(rows[r])) else None

    agg: list[tuple[str, str]] = []  # [(size, qty)]

    if max_col >= 2 and max_col >= max_row:
        # 세로형: 사이즈가 가장 많은 '왼쪽' 열을 사이즈 열로 본다.
        #   주의: 상의(A)·하의(C)가 둘 다 세로나열이면 두 열 개수가 동률이다.
        #   이때는 더 왼쪽(상의) 열을 채택한다(1차는 상의 사이즈 기준).
        best_cnt = max(col_counts.values())
        size_col = min(c for c, n in col_counts.items() if n == best_cnt)
        for r, c, s in size_cells:
            if c != size_col:
                continue
            # 수량은 '같은 행 바로 오른쪽 1칸'만 본다(아래 행 수량을 훔치면 안 됨).
            # 빈칸이면 그 사이즈 수량은 0 → 결과에서 제외(주문 안 된 사이즈).
            q = qty_of(get(r, c + 1))
            if q:
                agg.append((s, q))
    elif max_row >= 2:
        # 가로형: 사이즈가 가장 많은 행을 헤더로 보고, '바로 아래 1칸'에서 수량을 읽는다.
        size_row = max(row_counts, key=lambda k: row_counts[k])
        for r, c, s in size_cells:
            if r != size_row:
                continue
            q = qty_of(get(r + 1, c))
            if q:
                agg.append((s, q))
    else:
        # 분산형: 오른쪽 → 아래 순으로 수량 탐색.
        for r, c, s in size_cells:
            q = qty_of(get(r, c + 1)) or qty_of(get(r + 1, c))
            if q:
                agg.append((s, q))

    # 같은 사이즈가 여러 번 나오면 수량 합산.
    merged: dict[str, int] = {}
    for s, q in agg:
        merged[s] = merged.get(s, 0) + int(q)

    return [
        {"name": "", "number": "", "size": s, "qty": str(merged[s])}
        for s in merged
    ]


# ── 신양식③ 「STIZ 작성하기 신폼」 파서 ──────────────────────────────────────
# 수량 표기 "4장" / "4 장" / "4벌" 등에서 앞 숫자만 뽑는 정규식.
_QTY_NUM_PATTERN = re.compile(r"(\d+)")


def _parse_form_stiz(rows: list) -> list[dict]:
    """행 2차원 리스트에서 신양식③(STIZ '작성하기' 신폼)을 추출한다.

    양식③ 구조(실데이터 260213 연세대 추가주문서로 확정):
        헤더행:  순번 | 이니셜 | 배번 | 사이즈 | (빈) | 비고
        부제행:  (빈) | (빈)  | (빈) | 상의   | 하의 | (빈)   ← 건너뜀
        데이터:  1   | 이해솔 | 1   | M      | (빈) | 1장
      → 양식①과 달리 **맨 앞에 '순번' 컬럼**이 있고 **이름이 두 번째 열**이다.
        그래서 컬럼 위치를 헤더 키워드로 동적으로 찾는다(고정 오프셋 금지).

    반환: [{name=이니셜, number=배번, size=상의사이즈, qty=비고숫자}].
          헤더(순번+이니셜+사이즈)를 못 찾으면 빈 리스트(=양식③ 아님).
    """
    # 1) 헤더행 탐색: '순번' '이니셜' '사이즈' 가 한 행에 함께 있어야 양식③.
    #    (양식①은 '순번' 컬럼이 없어 여기서 걸러진다 → 두 양식 충돌 방지.)
    header_r = None
    col_seq = col_name = col_number = col_size = col_note = None
    for r in range(min(20, len(rows))):
        row = rows[r]
        cols: dict[str, int] = {}
        for c in range(min(10, len(row))):
            h = _norm_header(row[c])
            if h in _HEADER_SEQ and "seq" not in cols:
                cols["seq"] = c
            elif h in _HEADER_INITIAL and "name" not in cols:
                cols["name"] = c
            elif h in _HEADER_NUMBER and "number" not in cols:
                cols["number"] = c
            elif h in _HEADER_SIZE and "size" not in cols:
                cols["size"] = c
            elif h in _HEADER_NOTE and "note" not in cols:
                cols["note"] = c
        # 양식③ 식별 핵심: 순번 + 이니셜 + 사이즈가 모두 있는 행.
        if "seq" in cols and "name" in cols and "size" in cols:
            header_r = r
            col_seq = cols["seq"]
            col_name = cols["name"]
            col_number = cols.get("number", col_name + 1)  # 보통 이니셜 옆 칸.
            col_size = cols["size"]
            col_note = cols.get("note")  # 비고(수량). 없으면 qty="1".
            break
    if header_r is None:
        return []  # 양식③ 아님.

    # 2) 부제행(상의/하의) 건너뛰기: 헤더 바로 아래에 '상의'/'하의'가 보이면 1행 skip.
    start_r = header_r + 1
    if start_r < len(rows):
        sub = rows[start_r]
        joined = "".join(_norm_header(v) for v in sub[:8])
        if "상의" in joined or "하의" in joined:
            start_r += 1  # 부제행을 건너뛴다(상의 사이즈 = col_size 열 그대로).

    def cell(row, idx):
        return row[idx] if (idx is not None and idx < len(row)) else None

    # 3) 데이터 행 추출.
    result: list[dict] = []
    for r in range(start_r, len(rows)):
        row = rows[r]

        name = _to_str(cell(row, col_name))
        number = _to_str(cell(row, col_number))
        size = _normalize_size(cell(row, col_size))

        # 비고("4장")에서 수량 숫자만 추출. 없거나 못 읽으면 qty="1"(1선수 1행 기본).
        qty = "1"
        if col_note is not None:
            note_raw = _to_str(cell(row, col_note))
            m = _QTY_NUM_PATTERN.search(note_raw)
            if m:
                qty = m.group(1)

        # 핵심 결측행 skip: 이름·배번·사이즈가 모두 비면 데이터가 아니다.
        #   양식③ 끝은 순번만 미리 적혀 있고 나머지(이름/배번/사이즈)는 비어 있다
        #   (예: 39~100행). 이런 '순번만 있는 빈 행'을 건너뛴다(크래시·오염 방지).
        if not name and not number and not size:
            continue

        result.append({
            "name": name,
            "number": number,
            "size": size,   # 정규화 실패 시 ""(웹에서 강조 → 수기 보정)
            "qty": qty,     # 비고의 "4장" → "4". 없으면 "1".
        })

    return result


# ── 공개 함수 ────────────────────────────────────────────────────────────────
def parse_order(xlsx_path: str, warnings: list | None = None) -> list[dict]:
    """주문서 xlsx 를 읽어 표준 행 리스트로 반환한다.

    - 시트별로 양식①(이니셜 헤더) → 양식②(사이즈 집계) 순으로 파싱한 뒤,
      아래 **우선순위**로 '이번 주문' 시트를 채택한다(되돌림1 수정).
        1순위) 상단에 '주문번호'+'수량' 표식이 있는 시트 중 첫 번째(=이번 주문).
        2순위) 표식이 없으면 데이터가 나온 **첫 시트**(첫 시트가 항상 최신 접수분).
        3순위) 그래도 없으면 행을 가장 많이 찾은 시트(최후 폴백).
      → 과거 대량 접수분 시트를 '행 최다'라는 이유로 잘못 고르던 버그를 막는다.
    - 로드 실패/사이즈 없음 등은 빈 리스트 + 경고로 흘린다(절대 크래시하지 않음).

    반환: [{name, number, size, qty}] (값 전부 str, 실패 칸은 "")
    """
    warn = warnings if warnings is not None else []

    # data_only=True: 수식이 아닌 계산된 값을 읽는다(배번이 수식인 경우 대비).
    # read_only=True: 대용량/속도 최적화(시트를 메모리에 다 안 올림).
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        warn.append(f"⚠️ 주문서를 열 수 없습니다: {xlsx_path} ({e})")
        return []

    # 시트를 순회하며 (파싱결과, 표식여부)를 순서대로 모은다.
    # 데이터가 0행인 시트(표지/안내/폰트 시트)는 후보에서 제외한다.
    marked: list[dict] = []        # 1순위 후보: 표식 있고 데이터 있는 첫 시트
    first_nonempty: list[dict] = []  # 2순위 후보: 데이터 있는 첫 시트
    most_rows: list[dict] = []       # 3순위 후보: 행 최다 시트(최후 폴백)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 시트 전체 값을 2차원 리스트로 적재(read_only 모드 표준 방식).
            try:
                rows = [list(r) for r in ws.iter_rows(values_only=True)]
            except Exception as e:
                warn.append(f"⚠️ 시트 '{sheet_name}' 읽기 실패: {e}")
                continue
            if not rows:
                continue

            # 양식① 우선 → 비면 신양식③ → 비면 양식②.
            #   ①(이름=A열, 순번 없음)이 0행이면 ③(순번+이니셜 헤더)을 시도한다.
            #   ③은 '순번' 컬럼 유무로 ①과 구분되므로 충돌하지 않는다.
            parsed = _parse_form1(rows)
            if not parsed:
                parsed = _parse_form_stiz(rows)
            if not parsed:
                parsed = _parse_form2(rows)
            if not parsed:
                continue  # 데이터가 없는 시트는 후보에서 제외.

            # (1순위) '주문번호'+'수량' 표식이 있는 시트 — 가장 먼저 본 것만 기억.
            if not marked and _has_order_mark(rows):
                marked = parsed
            # (2순위) 데이터가 나온 첫 시트 — 가장 먼저 본 것만 기억.
            if not first_nonempty:
                first_nonempty = parsed
            # (3순위) 행 최다 시트 — 최후 폴백.
            if len(parsed) > len(most_rows):
                most_rows = parsed
    finally:
        wb.close()

    # 우선순위대로 채택: 표식 시트 > 첫 데이터 시트 > 행 최다 시트.
    best = marked or first_nonempty or most_rows

    if not best:
        warn.append(f"⚠️ 주문서에서 선수/사이즈 정보를 찾지 못했습니다: {xlsx_path}")

    # 사이즈가 비어 있는 행 수를 세어 안내(웹에서 붉게 강조할 대상).
    empty_size = sum(1 for row in best if not row["size"])
    if empty_size:
        warn.append(f"⚠️ 사이즈를 인식 못 한 행 {empty_size}개 — 확인이 필요합니다.")

    return best
